import csv
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles import Border, Side

thin = Side(style="thin")

border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)
      
#This block of code will create/format an Excel file and save to output folder 
workbook = Workbook()
sheet = workbook.active
sheet.title = "Sales Report"


#Define headers
headers = [
    "Order ID",
    "Date",
    "Salesperson",
    "Department",
    "Product",
    "Quantity",
    "Unit Price",
    "Total"

]


def calculate_total(quantity,Unit_price):
    return quantity * Unit_price

#get list length
max_item = len(headers)

#Loop to list
for r in range(1, max_item+1):
    print(headers[r-1])
    sheet.cell(5,r).value = headers[r-1]
    sheet.cell(5,r).font = Font(bold= True,color="FFFFFF")
    sheet.cell(5,r).fill = PatternFill(fill_type="solid",start_color="1F4E78")
    sheet.cell(5,r).alignment = Alignment(horizontal="center")


#Extract todays date with hours/mm/ss
generated_on = datetime.now().strftime("%d-%b-%Y %I:%M %p")


#Create Sheet description
sheet.cell(1,1).value = "Sales Report" 
sheet.cell(2,1).value ="Company"
sheet.cell(2,2).value = "ABC Corporation"
sheet.cell(3,1).value = "Generated On"
sheet.cell(3,2).value = generated_on

#Populate Data fields
counter = 6
with open("Input/sales.csv", newline="") as file:
    reader = csv.DictReader(file, delimiter="\t")
    for each in reader:
        sheet.cell(row = counter, column= 1).value = each['Order ID']
        sheet.cell(row = counter, column= 2).value = each['Date']
        sheet.cell(row = counter, column= 3).value = each['Salesperson']
        sheet.cell(row = counter, column= 4).value = each['Department']
        sheet.cell(row = counter, column= 5).value = each['Product']
        sheet.cell(row = counter, column= 6).value = each['Quantity']
        
        quantity = each['Quantity']
        unit_price = each['Unit Price']
        int_price = int(unit_price)
        sheet.cell(row = counter, column= 7).value = int_price
        sheet.cell(row = counter, column= 7).number_format ="₱#,##0.00"
        int_quantity = int(quantity)
        
        total = calculate_total(int_quantity,int_price)
        sheet.cell(row = counter, column= 8).value = total
        sheet.cell(row = counter, column= 8).number_format ="₱#,##0.00"

        counter = counter +1
        
        

#Format style
sheet.cell(1,1).font = Font(bold=True,size=16)


#Auto fit columns
for column in sheet.columns:
    max_length = 0
    for cell in column:
        if cell.value is not None:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value)) 
    column_letter = column[0].column_letter
    sheet.column_dimensions[column_letter].width = max_length+2


sheet.column_dimensions["G"].width += 4
sheet.column_dimensions["H"].width += 4

#Add border
for row in sheet.iter_rows(min_row=5, max_row=counter-1):
    for cell in row:
        cell.border = border



#Freeze column
sheet.freeze_panes = "A6"

#Merge title cells
sheet.merge_cells("A1:H1")
sheet["A1"].alignment = Alignment(horizontal="center")


filename = datetime.now().strftime("Output/Sales_Report_%Y%m%d_%H%M%S.xlsx")
workbook.save(filename)